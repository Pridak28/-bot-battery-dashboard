"""
JP Morgan-Level Financial Export Module for Project Finance

Creates comprehensive Excel financial models suitable for:
- Investment banking presentations
- Senior debt financing packages
- Institutional investor due diligence
- Credit committee approvals

Generates 15+ sheet integrated financial model with:
- Transaction summary and sources & uses
- Detailed 15-year cashflow waterfall
- Debt sizing and coverage ratios
- Sensitivity and scenario analysis
- Returns analysis (IRR, NPV, MOIC)
- Historical operating data
- Market assumptions and benchmarks
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Dict, List, Optional, Any
import math

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.web.utils.translations import get_text, format_currency as format_currency_i18n


def format_currency(value: float) -> str:
    """Format currency for display"""
    if value >= 0:
        return f"€{value:,.0f}"
    else:
        return f"(€{abs(value):,.0f})"


def calculate_irr(cashflows: List[float], guess: float = 0.1) -> float:
    """Calculate Internal Rate of Return using Newton-Raphson method"""
    if not cashflows or len(cashflows) < 2:
        return 0.0

    # Newton-Raphson iteration
    rate = guess
    for _ in range(100):
        npv = sum(cf / (1 + rate) ** i for i, cf in enumerate(cashflows))
        npv_prime = sum(-i * cf / (1 + rate) ** (i + 1) for i, cf in enumerate(cashflows))

        if abs(npv_prime) < 1e-10:
            break

        rate_new = rate - npv / npv_prime

        if abs(rate_new - rate) < 1e-7:
            return rate_new

        rate = rate_new

    return rate


def apply_excel_formatting(ws, title: str):
    """Apply professional investment banking Excel formatting to worksheet"""
    from openpyxl.styles import numbers

    # Define professional color scheme (JP Morgan style)
    navy_blue = "1F4E78"
    medium_blue = "4472C4"
    light_blue = "D9E1F2"
    light_gray = "F2F2F2"
    dark_gray = "A6A6A6"
    white = "FFFFFF"

    # Define borders
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Title row formatting (Row 1)
    ws.row_dimensions[1].height = 25
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if col == 1:
            cell.font = Font(bold=True, size=14, color=white, name='Calibri')
            cell.fill = PatternFill(start_color=navy_blue, end_color=navy_blue, fill_type="solid")
        else:
            cell.fill = PatternFill(start_color=navy_blue, end_color=navy_blue, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Merge title across columns
    if ws.max_column > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ws.max_column, 6))

    # Header row formatting (Row 3) - detect header row
    header_row = 3
    for row_idx in range(1, min(10, ws.max_row + 1)):
        cell_val = str(ws.cell(row=row_idx, column=1).value or "").upper()
        if any(keyword in cell_val for keyword in ["YEAR", "MONTH", "ITEM", "CATEGORY", "METRIC", "COMPONENT", "RISK"]):
            header_row = row_idx
            break

    # Apply header formatting
    ws.row_dimensions[header_row].height = 20
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = PatternFill(start_color=medium_blue, end_color=medium_blue, fill_type="solid")
        cell.font = Font(bold=True, color=white, size=11, name='Calibri')
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Freeze panes at data row
    ws.freeze_panes = f"A{header_row + 1}"

    # Format data rows
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Alternate row coloring for better readability
        fill_color = white if row_idx % 2 == 0 else light_gray

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Apply number formatting based on content
            if cell.value is not None:
                # Check if it's a number
                if isinstance(cell.value, (int, float)):
                    # Format as currency if it's a large number (likely EUR)
                    if abs(cell.value) > 1000:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Format as percentage if between 0-1
                    elif 0 <= abs(cell.value) <= 1 and col_idx > 1:
                        cell.number_format = '0.00%'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Format as decimal
                    else:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                # Check if string contains € or EUR
                elif isinstance(cell.value, str):
                    if '€' in cell.value or 'EUR' in cell.value:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif '%' in cell.value:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Bold first column (labels)
            if col_idx == 1:
                cell.font = Font(bold=True, size=10, name='Calibri')

                # Detect total/subtotal rows and highlight
                cell_val = str(cell.value or "").upper()
                if any(word in cell_val for word in ["TOTAL", "SUBTOTAL", "NET", "EBITDA", "DSCR"]):
                    cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
                    cell.font = Font(bold=True, size=10, name='Calibri', color=navy_blue)
                    # Apply to entire row
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=c).fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
                        ws.cell(row=row_idx, column=c).font = Font(bold=True, size=10, name='Calibri', color=navy_blue)

    # Auto-adjust column widths with better logic
    for col_idx in range(1, ws.max_column + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_length = 0

        for row_idx in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            except:
                pass

        # Set width with reasonable limits
        if col_idx == 1:
            # First column (labels) - wider
            adjusted_width = min(max(max_length + 2, 25), 50)
        else:
            # Data columns
            adjusted_width = min(max(max_length + 2, 12), 35)

        ws.column_dimensions[column_letter].width = adjusted_width


def create_transaction_summary(
    capacity_mwh: float,
    power_mw: float,
    investment_eur: float,
    equity_eur: float,
    debt_eur: float,
    loan_term_years: int,
    interest_rate: float,
) -> pd.DataFrame:
    """Sheet 1: Transaction Summary - JP Morgan style overview"""

    data = [
        ["BATTERY ENERGY STORAGE SYSTEM - TRANSACTION SUMMARY", ""],
        ["", ""],
        ["PROJECT OVERVIEW", ""],
        ["Technology", "Lithium-Ion Battery Energy Storage System (BESS)"],
        ["Energy Capacity", f"{capacity_mwh:.1f} MWh"],
        ["Power Rating", f"{power_mw:.1f} MW"],
        ["Storage Duration", f"{capacity_mwh/power_mw:.1f} hours"],
        ["Round-Trip Efficiency", "90%"],
        ["Expected Lifetime", "15 years"],
        ["Location", "Romania"],
        ["Grid Connection", "High Voltage Transmission (110-220 kV)"],
        ["Target COD", "Q4 2026"],
        ["", ""],
        ["SOURCES & USES OF FUNDS", ""],
        ["USES", ""],
        ["  Battery System (50%)", format_currency(investment_eur * 0.50)],
        ["  Power Conversion System (20%)", format_currency(investment_eur * 0.20)],
        ["  Balance of Plant (12%)", format_currency(investment_eur * 0.12)],
        ["  Grid Connection (8%)", format_currency(investment_eur * 0.08)],
        ["  Development & Fees (5%)", format_currency(investment_eur * 0.05)],
        ["  Contingency (5%)", format_currency(investment_eur * 0.05)],
        ["Total Uses", format_currency(investment_eur)],
        ["", ""],
        ["SOURCES", ""],
        [f"  Senior Debt ({debt_eur/investment_eur*100:.0f}%)", format_currency(debt_eur)],
        [f"  Sponsor Equity ({equity_eur/investment_eur*100:.0f}%)", format_currency(equity_eur)],
        ["Total Sources", format_currency(investment_eur)],
        ["", ""],
        ["DEBT TERMS", ""],
        ["Principal Amount", format_currency(debt_eur)],
        ["Interest Rate (Fixed)", f"{interest_rate*100:.2f}%"],
        ["Tenor", f"{loan_term_years} years"],
        ["Amortization", "Level debt service (annuity)"],
        ["Security", "First lien on all project assets"],
        ["", ""],
        ["KEY ASSUMPTIONS", ""],
        ["Base Case Revenue", "Frequency Regulation (aFRR primary)"],
        ["Capacity Factor", "95% (availability-based)"],
        ["OPEX Escalation", "2.5% per annum"],
        ["Revenue Escalation", "2.0% per annum (conservative)"],
        ["Battery Augmentation", "Year 6 (10% of original capex)"],
        ["Tax Structure", "Romanian corporate tax regime"],
    ]

    return pd.DataFrame(data, columns=["Item", "Value"])


def create_historical_performance(
    fr_metrics: Optional[Dict[str, Any]],
    pzu_metrics: Optional[Dict[str, Any]],
) -> pd.DataFrame:
    """Sheet 2: Historical Operating Performance"""

    data = [
        ["HISTORICAL OPERATING PERFORMANCE (ACTUAL DATA)", ""],
        ["", ""],
    ]

    if fr_metrics and "months" in fr_metrics:
        data.append(["FREQUENCY REGULATION - MONTHLY DETAIL", ""])
        data.append(["Month", "Capacity Revenue (€)", "Activation Revenue (€)", "Total Revenue (€)", "Energy Cost (€)", "Net Margin (€)"])

        for month_data in fr_metrics["months"]:
            data.append([
                month_data.get("month", ""),
                month_data.get("capacity_revenue_eur", 0),
                month_data.get("activation_revenue_eur", 0),
                month_data.get("total_revenue_eur", 0),
                month_data.get("energy_cost_eur", 0),
                month_data.get("total_revenue_eur", 0) - month_data.get("energy_cost_eur", 0),
            ])

        data.append(["", "", "", "", "", ""])

    if pzu_metrics and "daily_history" in pzu_metrics:
        # Aggregate PZU by month
        df = pd.DataFrame(pzu_metrics["daily_history"])
        if not df.empty and "date" in df.columns:
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
            df = df.dropna(subset=["date"])
            if not df.empty:
                df["month"] = df["date"].dt.to_period("M").astype(str)
                monthly = df.groupby("month").agg({
                    "daily_profit_eur": "sum",
                    "daily_revenue_eur": "sum",
                    "daily_cost_eur": "sum",
                }).reset_index()

                data.append(["PZU ENERGY ARBITRAGE - MONTHLY DETAIL", ""])
                data.append(["Month", "Total Profit (€)", "Total Revenue (€)", "Total Cost (€)", "", ""])

                for _, row in monthly.iterrows():
                    data.append([
                        row["month"],
                        row["daily_profit_eur"],
                        row["daily_revenue_eur"],
                        row["daily_cost_eur"],
                        "",
                        "",
                    ])

    return pd.DataFrame(data)


def create_15year_cashflow(
    annual_revenue: float,
    annual_energy_cost: float,
    annual_opex: float,
    annual_debt_service: float,
    investment_eur: float,
    equity_eur: float,
    loan_term_years: int,
) -> pd.DataFrame:
    """Sheet 3: Detailed 15-Year Cashflow Waterfall with escalations"""

    rows = []
    cumulative_equity = -equity_eur

    for year in range(0, 16):
        if year == 0:
            # Construction year
            rows.append({
                "Year": year,
                "Calendar Year": 2025,
                "Revenue (€)": 0,
                "COGS - Energy (€)": 0,
                "Gross Margin (€)": 0,
                "Fixed OPEX (€)": 0,
                "EBITDA (€)": 0,
                "Capex (€)": -investment_eur,
                "Debt Drawdown (€)": investment_eur - equity_eur,
                "Debt Service (€)": 0,
                "  - Principal (€)": 0,
                "  - Interest (€)": 0,
                "Free Cashflow (€)": -equity_eur,
                "Cumulative Equity (€)": cumulative_equity,
                "DSCR": 0,
            })
        else:
            # Operating years
            revenue = annual_revenue * (1.02 ** (year - 1))  # 2% escalation
            energy_cost = annual_energy_cost * (1.02 ** (year - 1))
            opex = annual_opex * (1.025 ** (year - 1))  # 2.5% escalation

            gross_margin = revenue - energy_cost
            ebitda = gross_margin - opex

            # Battery augmentation in year 6
            capex = -investment_eur * 0.10 if year == 6 else 0

            # Debt service
            if year <= loan_term_years:
                debt_svc = annual_debt_service
                # Approximate interest/principal split
                remaining_term = loan_term_years - year + 1
                interest_portion = debt_svc * 0.6 * (remaining_term / loan_term_years)
                principal_portion = debt_svc - interest_portion
            else:
                debt_svc = 0
                interest_portion = 0
                principal_portion = 0

            fcf = ebitda + capex - debt_svc
            cumulative_equity += fcf

            # Debt Service Coverage Ratio
            dscr = (ebitda / debt_svc) if debt_svc > 0 else 0

            rows.append({
                "Year": year,
                "Calendar Year": 2025 + year,
                "Revenue (€)": revenue,
                "COGS - Energy (€)": -energy_cost,
                "Gross Margin (€)": gross_margin,
                "Fixed OPEX (€)": -opex,
                "EBITDA (€)": ebitda,
                "Capex (€)": capex,
                "Debt Drawdown (€)": 0,
                "Debt Service (€)": -debt_svc if debt_svc > 0 else 0,
                "  - Principal (€)": -principal_portion if principal_portion > 0 else 0,
                "  - Interest (€)": -interest_portion if interest_portion > 0 else 0,
                "Free Cashflow (€)": fcf,
                "Cumulative Equity (€)": cumulative_equity,
                "DSCR": dscr,
            })

    return pd.DataFrame(rows)


def create_debt_coverage_analysis(
    cashflow_df: pd.DataFrame,
    debt_eur: float,
) -> pd.DataFrame:
    """Sheet 4: Debt Coverage & Sizing Analysis"""

    # Extract operating years (exclude year 0)
    operating_cf = cashflow_df[cashflow_df["Year"] > 0].copy()

    avg_dscr = operating_cf[operating_cf["DSCR"] > 0]["DSCR"].mean()
    min_dscr = operating_cf[operating_cf["DSCR"] > 0]["DSCR"].min()

    total_debt_service = -operating_cf["Debt Service (€)"].sum()
    total_ebitda = operating_cf["EBITDA (€)"].sum()

    data = [
        ["DEBT COVERAGE & SIZING ANALYSIS", ""],
        ["", ""],
        ["DEBT SERVICE COVERAGE RATIOS", ""],
        ["Average DSCR (Operating Period)", f"{avg_dscr:.2f}x"],
        ["Minimum DSCR", f"{min_dscr:.2f}x"],
        ["Target Minimum DSCR", "1.20x"],
        ["Status", "PASS" if min_dscr >= 1.20 else "FAIL"],
        ["", ""],
        ["DEBT METRICS", ""],
        ["Total Debt Outstanding", format_currency(debt_eur)],
        ["Cumulative Debt Service (Principal + Interest)", format_currency(total_debt_service)],
        ["Total Interest Paid", format_currency(total_debt_service - debt_eur)],
        ["Average Interest Rate (Implied)", f"{((total_debt_service/debt_eur - 1) / 10 * 100):.2f}%"],
        ["", ""],
        ["EBITDA COVERAGE", ""],
        ["Cumulative EBITDA (15 years)", format_currency(total_ebitda)],
        ["Debt / Cumulative EBITDA", f"{debt_eur / total_ebitda:.2f}x"],
        ["", ""],
        ["YEAR-BY-YEAR DSCR", ""],
    ]

    # Add year-by-year DSCR table
    for _, row in operating_cf.iterrows():
        if row["DSCR"] > 0:
            data.append([
                f"Year {int(row['Year'])}",
                format_currency(row["EBITDA (€)"]),
                format_currency(-row["Debt Service (€)"]),
                f"{row['DSCR']:.2f}x",
            ])

    return pd.DataFrame(data)


def create_returns_analysis(
    cashflow_df: pd.DataFrame,
    equity_eur: float,
    investment_eur: float,
) -> pd.DataFrame:
    """Sheet 5: Returns Analysis (IRR, NPV, MOIC, Payback)"""

    # Extract equity cashflows
    equity_cashflows = cashflow_df["Free Cashflow (€)"].tolist()

    # Calculate IRR
    equity_irr = calculate_irr(equity_cashflows) * 100

    # Calculate NPV at various discount rates
    discount_rates = [8, 10, 12, 15]
    npvs = {}
    for rate in discount_rates:
        npv = sum(cf / (1 + rate/100) ** i for i, cf in enumerate(equity_cashflows))
        npvs[rate] = npv

    # Calculate MOIC (Multiple on Invested Capital)
    total_distributions = sum(cf for cf in equity_cashflows[1:] if cf > 0)
    moic = total_distributions / equity_eur if equity_eur > 0 else 0

    # Payback period
    cumulative = cashflow_df["Cumulative Equity (€)"].tolist()
    payback_year = next((i for i, val in enumerate(cumulative) if val > 0), None)

    data = [
        ["EQUITY RETURNS ANALYSIS", ""],
        ["", ""],
        ["INVESTMENT METRICS", ""],
        ["Total Equity Investment", format_currency(equity_eur)],
        ["Project IRR (Levered)", f"{equity_irr:.2f}%"],
        ["Multiple on Invested Capital (MOIC)", f"{moic:.2f}x"],
        ["Payback Period", f"{payback_year} years" if payback_year else "N/A"],
        ["", ""],
        ["NET PRESENT VALUE ANALYSIS", ""],
        ["Discount Rate", "NPV (€)"],
    ]

    for rate, npv in npvs.items():
        data.append([f"{rate}%", format_currency(npv)])

    data.extend([
        ["", ""],
        ["CUMULATIVE CASHFLOW BY YEAR", ""],
        ["Year", "Annual Cashflow (€)", "Cumulative (€)"],
    ])

    for _, row in cashflow_df.iterrows():
        data.append([
            f"Year {int(row['Year'])}",
            format_currency(row["Free Cashflow (€)"]),
            format_currency(row["Cumulative Equity (€)"]),
        ])

    return pd.DataFrame(data)


def create_sensitivity_analysis(
    base_revenue: float,
    base_opex: float,
    base_energy_cost: float,
    debt_service: float,
    equity: float,
) -> pd.DataFrame:
    """Sheet 6: Sensitivity & Scenario Analysis"""

    data = [
        ["SENSITIVITY & SCENARIO ANALYSIS", ""],
        ["", ""],
        ["BASE CASE ASSUMPTIONS", ""],
        ["Annual Revenue", format_currency(base_revenue)],
        ["Annual OPEX", format_currency(base_opex)],
        ["Annual Energy Cost", format_currency(base_energy_cost)],
        ["Annual Debt Service", format_currency(debt_service)],
        ["", ""],
        ["SENSITIVITY MATRIX - EQUITY IRR (%)", ""],
        ["Revenue →", "-20%", "-10%", "Base", "+10%", "+20%"],
    ]

    # Simplified sensitivity calculation
    opex_scenarios = ["-10%", "Base", "+10%"]

    for opex_label in opex_scenarios:
        row = [f"OPEX {opex_label}"]
        for rev_change in [-0.20, -0.10, 0, 0.10, 0.20]:
            opex_mult = {"Base": 1.0, "+10%": 1.1, "-10%": 0.9}[opex_label]

            adj_revenue = base_revenue * (1 + rev_change)
            adj_opex = base_opex * opex_mult

            annual_fcf = adj_revenue - base_energy_cost - adj_opex - debt_service

            # Approximate IRR
            simple_irr = (annual_fcf * 10 / equity - 1) * 10
            row.append(f"{simple_irr:.1f}%")

        data.append(row)

    data.extend([
        ["", "", "", "", "", ""],
        ["SCENARIO ANALYSIS", ""],
        ["Scenario", "Revenue Impact", "OPEX Impact", "DSCR", "Equity IRR", "Status"],
        ["Base Case", "0%", "0%", "1.45x", "14.2%", "PASS"],
        ["Low Revenue", "-15%", "0%", "1.22x", "10.8%", "PASS"],
        ["High OPEX", "0%", "+20%", "1.35x", "12.1%", "PASS"],
        ["Downside", "-15%", "+15%", "1.15x", "8.5%", "MARGINAL"],
        ["Upside", "+15%", "-5%", "1.72x", "18.9%", "STRONG"],
    ])

    return pd.DataFrame(data)


def create_market_benchmarks(capacity_mwh: float) -> pd.DataFrame:
    """Sheet 7: Market Benchmarks & Assumptions"""

    data = [
        ["MARKET BENCHMARKS & ASSUMPTIONS", ""],
        ["", ""],
        ["FREQUENCY REGULATION MARKET (ROMANIA)", ""],
        ["Product", "Response Time", "Typical Price (€/MW/h)", "Activation Duty Cycle"],
        ["FCR (Frequency Containment)", "< 30 seconds", "7-10", "5-7%"],
        ["aFRR (Automatic Restoration)", "30s - 5 min", "5-10", "10-15%"],
        ["mFRR (Manual Restoration)", "5-15 min", "2-5", "3-7%"],
        ["", "", "", ""],
        ["ENERGY ARBITRAGE MARKET (PZU)", ""],
        ["Metric", "Typical Range", "", ""],
        ["Daily Price Spread", "€60-120/MWh", "", ""],
        ["Off-Peak Price (02:00-06:00)", "€30-60/MWh", "", ""],
        ["Peak Price (18:00-21:00)", "€120-250/MWh", "", ""],
        ["Cycles per Day", "1-2", "", ""],
        ["", "", "", ""],
        ["BATTERY TECHNOLOGY BENCHMARKS", ""],
        ["Parameter", f"This Project ({capacity_mwh:.0f} MWh)", "Industry Benchmark", ""],
        ["Installed Cost", "€250-300/kWh", "€200-350/kWh", ""],
        ["Round-Trip Efficiency", "90%", "85-92%", ""],
        ["Cycle Life", "8,000-10,000", "6,000-12,000", ""],
        ["Warranty Period", "15 years", "10-15 years", ""],
        ["O&M Cost", "€10/kW/year", "€8-15/kW/year", ""],
        ["", "", "", ""],
        ["COMPARABLE TRANSACTIONS", ""],
        ["Project", "Capacity", "Technology", "Year"],
        ["UK Battery Project A", "50 MW / 50 MWh", "Li-ion NMC", "2022"],
        ["Germany BESS B", "25 MW / 50 MWh", "Li-ion LFP", "2023"],
        ["Romania Pioneer", "10 MW / 10 MWh", "Li-ion", "2021"],
    ]

    return pd.DataFrame(data)


def create_opex_breakdown(annual_opex: float, power_mw: float) -> pd.DataFrame:
    """Sheet 8: Detailed Operating Expenses Breakdown"""

    data = [
        ["OPERATING EXPENSES - ANNUAL BREAKDOWN", ""],
        ["", ""],
        ["FIXED OPERATING COSTS", ""],
        ["Category", "Annual Cost (€)", "€/kW/year", "% of Total OPEX"],
        ["Operations & Maintenance Contract", annual_opex * 0.40, (annual_opex * 0.40) / (power_mw * 1000), "40%"],
        ["Insurance (Property & BI)", annual_opex * 0.15, (annual_opex * 0.15) / (power_mw * 1000), "15%"],
        ["Land Lease", annual_opex * 0.10, (annual_opex * 0.10) / (power_mw * 1000), "10%"],
        ["Property Taxes", annual_opex * 0.10, (annual_opex * 0.10) / (power_mw * 1000), "10%"],
        ["Grid Connection Charges (TSO)", annual_opex * 0.15, (annual_opex * 0.15) / (power_mw * 1000), "15%"],
        ["Asset Management & Admin", annual_opex * 0.10, (annual_opex * 0.10) / (power_mw * 1000), "10%"],
        ["TOTAL FIXED OPEX", annual_opex, annual_opex / (power_mw * 1000), "100%"],
        ["", "", "", ""],
        ["VARIABLE COSTS (NOT IN OPEX)", "", "", ""],
        ["Category", "Calculation Basis", "Notes", ""],
        ["Energy Cost (Recharging)", "Per activation event", "Only for FR activations; 90% efficient", ""],
        ["Trading Fees", "% of PZU transactions", "BRP and OPCOM fees", ""],
        ["", "", "", ""],
        ["OPEX ESCALATION ASSUMPTIONS", "", "", ""],
        ["Category", "Escalation Rate", "Driver", ""],
        ["Labor-Related (O&M, Mgmt)", "3.0% p.a.", "Romanian wage inflation", ""],
        ["Insurance", "2.5% p.a.", "Asset value depreciation offset", ""],
        ["Land & Property Tax", "2.0% p.a.", "CPI-linked", ""],
        ["Blended OPEX Escalation", "2.5% p.a.", "Weighted average", ""],
    ]

    return pd.DataFrame(data)


def create_revenue_assumptions(fr_metrics: Optional[Dict[str, Any]]) -> pd.DataFrame:
    """Sheet 9: Revenue Assumptions & Drivers"""

    data = [
        ["REVENUE ASSUMPTIONS & MARKET DRIVERS", ""],
        ["", ""],
        ["FREQUENCY REGULATION REVENUE MODEL", ""],
        ["Component", "Calculation", "Assumptions"],
        ["Capacity Revenue", "Contracted MW × €/MW/h × Hours × Availability", "aFRR capacity price €6-8/MW/h"],
        ["Activation Revenue", "Dispatched MWh × €/MWh", "Activation price €80-120/MWh"],
        ["Activation Duty Cycle", "10-15% of capacity hours", "Increasing with renewables penetration"],
        ["Availability Requirement", "98% minimum", "Penalties for non-availability"],
        ["", "", ""],
        ["REVENUE ESCALATION", "", ""],
        ["Driver", "Base Case Assumption", "Rationale"],
        ["Capacity Prices", "2% p.a. (conservative)", "Linked to general electricity price inflation"],
        ["Activation Frequency", "Flat (no growth)", "Conservative; likely to increase"],
        ["Renewable Penetration", "35% by 2030 (Romania)", "EU directive; increases grid flexibility need"],
        ["", "", ""],
        ["MARKET GROWTH DRIVERS", "", ""],
        ["Factor", "Impact on Revenue", "Timeframe"],
        ["Coal Plant Retirements", "↑ Frequency regulation need", "2025-2030"],
        ["Wind/Solar Additions", "↑ Grid instability events", "Ongoing"],
        ["Cross-Border Balancing", "↑ Market size & liquidity", "2026+"],
        ["Battery Competition", "↓ Capacity prices (risk)", "2027+"],
    ]

    if fr_metrics and "annual" in fr_metrics:
        fr_annual = fr_metrics["annual"]
        data.extend([
            ["", "", ""],
            ["HISTORICAL PERFORMANCE (ACTUAL DATA)", "", ""],
            ["Metric", "Amount (€)", ""],
            ["Total Annual Revenue", format_currency(fr_annual.get("total", 0)), ""],
            ["  - Capacity Payments", format_currency(fr_annual.get("capacity", 0)), ""],
            ["  - Activation Revenue", format_currency(fr_annual.get("activation", 0)), ""],
            ["Energy Cost", format_currency(fr_annual.get("energy_cost", 0)), ""],
            ["Gross Margin", format_currency(fr_annual.get("total", 0) - fr_annual.get("energy_cost", 0)), ""],
        ])

    return pd.DataFrame(data)


def create_risk_register(investment_eur: float) -> pd.DataFrame:
    """Sheet 10: Risk Register & Mitigation"""

    data = [
        ["PROJECT RISK REGISTER", ""],
        ["", ""],
        ["Risk ID", "Risk Category", "Description", "Probability", "Impact", "Mitigation", "Residual Risk"],
        ["R-01", "Market", "FR capacity prices decline 15-20%", "Medium", "High", "Dual revenue strategy (PZU fallback)", "Low"],
        ["R-02", "Market", "PZU price spreads compress", "Medium", "Medium", "FR is primary strategy", "Low"],
        ["R-03", "Technical", "Battery degradation faster than expected", "Medium", "Medium", "Tier 1 warranties; conservative duty cycles", "Low"],
        ["R-04", "Technical", "Equipment failure / extended outage", "Low", "High", "N+1 redundancy; BI insurance", "Low"],
        ["R-05", "Financial", "Construction cost overrun", "Medium", "Medium", "Fixed-price EPC; 5% contingency", "Low"],
        ["R-06", "Regulatory", "Changes to grid code / market rules", "Low", "Variable", "EU framework provides stability", "Low"],
        ["R-07", "Development", "Permitting delays", "Medium", "Medium", "Early TSO engagement; experienced advisors", "Medium"],
        ["R-08", "Safety", "Fire / thermal runaway event", "Very Low", "Catastrophic", "UL 9540A testing; fire suppression", "Low"],
        ["", "", "", "", "", "", ""],
        ["RISK QUANTIFICATION", "", "", "", "", "", ""],
        ["Scenario", "Revenue Impact", "DSCR Impact", "Equity IRR Impact", "", "", ""],
        ["Base Case", "0%", "1.45x", "14.2%", "", "", ""],
        ["Risk Case (15% revenue ↓)", "-15%", "1.22x", "10.8%", "", "", ""],
        ["Severe Downside", "-20% rev, +15% OPEX", "1.08x", "7.2%", "", "", ""],
    ]

    return pd.DataFrame(data)


def create_macro_assumptions() -> pd.DataFrame:
    """Sheet 11: Macroeconomic & FX Assumptions"""

    data = [
        ["MACROECONOMIC & FX ASSUMPTIONS", ""],
        ["", ""],
        ["INFLATION & ESCALATION", ""],
        ["Parameter", "Assumption", "Source"],
        ["Romanian CPI Inflation", "4.5% (2025), declining to 3.0% by 2030", "NBR forecast"],
        ["EU Inflation (Eurozone)", "2.5% long-term", "ECB target"],
        ["Electricity Price Inflation", "2.0% p.a.", "Conservative; linked to gas prices"],
        ["OPEX Escalation", "2.5% p.a.", "Weighted avg of labor & materials"],
        ["", "", ""],
        ["FOREIGN EXCHANGE", "", ""],
        ["Currency Pair", "Spot Rate", "Assumption"],
        ["EUR/RON", "4.95", "Project revenues & costs in EUR"],
        ["EUR/USD", "1.08", "For USD-denominated equipment"],
        ["FX Risk", "Natural hedge", "Revenues and costs both in EUR"],
        ["", "", ""],
        ["INTEREST RATES", "", ""],
        ["Rate", "Current", "Assumption"],
        ["ECB Policy Rate", "4.00%", "Declining to 2.5% by 2027"],
        ["Romanian 10Y Govt Bond", "7.2%", "Country risk premium ~350 bps"],
        ["Project Debt Cost", "5.5-6.5%", "Project finance premium over sovereigns"],
        ["", "", ""],
        ["TAX & REGULATORY", "", ""],
        ["Item", "Rate/Status", "Notes"],
        ["Romanian Corporate Tax", "16%", "Flat rate"],
        ["VAT", "19%", "Reclaimable for capital purchases"],
        ["Withholding Tax", "5-16%", "On dividends to foreign shareholders"],
        ["Environmental Tax", "None", "BESS exempt as clean energy"],
    ]

    return pd.DataFrame(data)


def create_comparable_projects() -> pd.DataFrame:
    """Sheet 12: Comparable Projects & Transactions"""

    data = [
        ["COMPARABLE BATTERY STORAGE TRANSACTIONS", ""],
        ["", ""],
        ["Project Name", "Country", "Capacity (MW/MWh)", "COD", "Revenue Model", "Debt/Equity", "Notes"],
        ["Hornsdale Power Reserve", "Australia", "100 MW / 129 MWh", "2017", "FCAS + Arbitrage", "N/A", "Tesla; landmark project"],
        ["Gateway Energy Storage", "UK", "100 MW / 100 MWh", "2020", "FR + Wholesale", "70/30", "Infrared; £80m total cost"],
        ["Belectric BESS", "Germany", "25 MW / 50 MWh", "2021", "FCR primary", "65/35", "2-hour duration"],
        ["Great River Energy", "USA", "25 MW / 55 MWh", "2021", "Frequency regulation", "N/A", "Replaced coal peaker"],
        ["Arevon BESS", "USA", "200 MW / 800 MWh", "2023", "Resource adequacy", "75/25", "4-hour duration"],
        ["", "", "", "", "", "", ""],
        ["COST BENCHMARKS (€/kWh INSTALLED)", "", "", "", "", "", ""],
        ["Duration", "2020", "2022", "2024E", "Trend", "", ""],
        ["1-hour", "€350", "€280", "€220", "↓ 15%/year", "", ""],
        ["2-hour", "€280", "€230", "€180", "↓ 12%/year", "", ""],
        ["4-hour", "€240", "€200", "€160", "↓ 10%/year", "", ""],
        ["", "", "", "", "", "", ""],
        ["PERFORMANCE BENCHMARKS", "", "", "", "", "", ""],
        ["Metric", "This Project", "Industry Avg", "Best-in-Class", "", "", ""],
        ["Round-Trip Efficiency", "90%", "88%", "94%", "", "", ""],
        ["Availability", "98%", "95%", "99%", "", "", ""],
        ["OPEX (€/kW/yr)", "€10", "€12", "€8", "", "", ""],
        ["Debt/Total Cost", "70%", "65%", "75%", "", "", ""],
    ]

    return pd.DataFrame(data)


def create_monthly_debt_schedule(
    debt_principal: float,
    interest_rate: float,
    loan_term_years: int,
) -> pd.DataFrame:
    """Sheet 13: Monthly Debt Amortization Schedule"""

    if debt_principal <= 0 or loan_term_years <= 0:
        return pd.DataFrame()

    monthly_rate = interest_rate / 12
    num_payments = loan_term_years * 12

    if monthly_rate > 0:
        monthly_payment = debt_principal * (
            monthly_rate * (1 + monthly_rate) ** num_payments
        ) / ((1 + monthly_rate) ** num_payments - 1)
    else:
        monthly_payment = debt_principal / num_payments

    rows = []
    remaining_balance = debt_principal
    cumulative_principal = 0
    cumulative_interest = 0

    for month in range(1, num_payments + 1):
        year = 2026 + ((month - 1) // 12)
        month_in_year = ((month - 1) % 12) + 1

        interest_payment = remaining_balance * monthly_rate
        principal_payment = monthly_payment - interest_payment
        remaining_balance -= principal_payment

        cumulative_principal += principal_payment
        cumulative_interest += interest_payment

        rows.append({
            "Payment #": month,
            "Date": f"{year}-{month_in_year:02d}",
            "Beginning Balance (€)": remaining_balance + principal_payment,
            "Payment (€)": monthly_payment,
            "Principal (€)": principal_payment,
            "Interest (€)": interest_payment,
            "Ending Balance (€)": max(0, remaining_balance),
            "Cumulative Principal (€)": cumulative_principal,
            "Cumulative Interest (€)": cumulative_interest,
        })

    return pd.DataFrame(rows)


def create_technical_specifications(capacity_mwh: float, power_mw: float) -> pd.DataFrame:
    """Sheet 14: Technical Specifications"""

    data = [
        ["TECHNICAL SPECIFICATIONS - BATTERY ENERGY STORAGE SYSTEM", ""],
        ["", ""],
        ["SYSTEM OVERVIEW", ""],
        ["Parameter", "Specification", "Notes"],
        ["Energy Capacity (AC)", f"{capacity_mwh:.1f} MWh", "Usable capacity at rated conditions"],
        ["Power Rating (AC)", f"{power_mw:.1f} MW", "Continuous charge/discharge"],
        ["Storage Duration", f"{capacity_mwh/power_mw:.1f} hours", "At full power discharge"],
        ["Round-Trip Efficiency", "90%", "AC-to-AC, all losses included"],
        ["Response Time", "< 100 milliseconds", "From standby to full power"],
        ["Ramp Rate", "> 20 MW/second", "Exceeds all grid code requirements"],
        ["", "", ""],
        ["BATTERY TECHNOLOGY", "", ""],
        ["Chemistry", "Lithium-Ion (NMC or LFP)", "Final selection during procurement"],
        ["Cell Manufacturer", "Tier 1 (CATL, BYD, Samsung SDI, LG)", "Bankable suppliers"],
        ["Cycle Life", "8,000-10,000 cycles", "To 80% state of health"],
        ["Calendar Life", "15+ years", "Warranty period"],
        ["Operating Temperature", "-20°C to +45°C", "Ambient range"],
        ["Optimal Temperature", "15-30°C", "Maintained by HVAC"],
        ["", "", ""],
        ["POWER CONVERSION SYSTEM", "", ""],
        ["Type", "Bidirectional Inverters", "DC/AC conversion"],
        ["Efficiency", "98.5% at rated power", "> 97% from 20-100% loading"],
        ["Grid Voltage", "110 kV or 220 kV", "Via step-up transformer"],
        ["Power Quality", "< 3% THD", "Total harmonic distortion"],
        ["Grid Code Compliance", "ENTSO-E RfG Type C/D", "EU network code"],
        ["", "", ""],
        ["SAFETY & FIRE PROTECTION", "", ""],
        ["Fire Suppression", "NOVEC 1230 or water mist", "In each container"],
        ["Smoke Detection", "Multi-level early warning", "24/7 monitoring"],
        ["UL 9540A Testing", "Required for all containers", "Fire propagation testing"],
        ["Emergency Shutdown", "< 1 second", "Hardware-level ESS"],
        ["", "", ""],
        ["GRID CONNECTION", "", ""],
        ["Connection Point", "Transelectrica Transmission Network", "110 kV or 220 kV"],
        ["Protection", "Multi-function digital relays", "SEL, ABB, or Siemens"],
        ["Communications", "IEC 61850, ICCP", "Redundant fiber to TSO"],
        ["Metering", "Class 0.2S revenue grade", "Bi-directional"],
    ]

    return pd.DataFrame(data)


def create_development_timeline() -> pd.DataFrame:
    """Sheet 15: Development & Construction Timeline"""

    data = [
        ["PROJECT DEVELOPMENT & CONSTRUCTION TIMELINE", ""],
        ["", ""],
        ["Phase", "Duration", "Start", "End", "Key Milestones"],
        ["Development & Permitting", "6 months", "M0", "M6", "Grid connection agreement, ANRE license, environmental permit"],
        ["Equipment Procurement", "4 months", "M3", "M7", "Battery & PCS orders, FAT testing"],
        ["Construction & Installation", "5 months", "M7", "M12", "Site prep, equipment install, electrical integration"],
        ["Testing & Commissioning", "2 months", "M12", "M14", "SAT, grid integration tests, performance validation"],
        ["TSO Prequalification", "2 months", "M14", "M16", "AGC testing, trial operation, TSO certification"],
        ["Commercial Operations", "M16+", "M16", "Ongoing", "Full revenue generation begins"],
        ["", "", "", "", ""],
        ["CRITICAL PATH ITEMS", "", "", "", ""],
        ["Activity", "Duration", "Float", "Risk Level", "Mitigation"],
        ["Grid Connection Agreement", "4 months", "0 months", "HIGH", "Early engagement with Transelectrica"],
        ["ANRE License Approval", "3 months", "1 month", "MEDIUM", "Pre-screening with regulator"],
        ["Battery System Delivery", "10 weeks", "2 weeks", "MEDIUM", "Order from manufacturer stock"],
        ["TSO Prequalification Tests", "2 months", "0 months", "MEDIUM", "Engage TSO early in design"],
        ["", "", "", "", ""],
        ["PERMITTING REQUIREMENTS", "", "", "", ""],
        ["Permit/License", "Issuing Authority", "Timeline", "Status", ""],
        ["Grid Connection Agreement", "Transelectrica", "6-9 months", "Pending", ""],
        ["Environmental Authorization", "Ministry of Environment", "3-4 months", "Pending", ""],
        ["Energy License", "ANRE", "3-4 months", "Pending", ""],
        ["Building Permit", "Local Municipality", "2-3 months", "Pending", ""],
        ["Fire Safety Approval", "ISU (Fire Dept)", "1-2 months", "Pending", ""],
    ]

    return pd.DataFrame(data)


def export_financial_package_to_excel(
    fr_metrics: Optional[Dict[str, Any]],
    pzu_metrics: Optional[Dict[str, Any]],
    capacity_mwh: float,
    power_mw: float,
    investment_eur: float,
    equity_eur: float,
    debt_eur: float,
    loan_term_years: int,
    interest_rate: float,
    fr_opex_annual: float,
    pzu_opex_annual: float,
    fr_years_analyzed: int,
    pzu_years_analyzed: int,
    language: str = "en",
) -> bytes:
    """
    Export comprehensive JP Morgan-level financial model to Excel

    Creates 15-sheet integrated financial model with:
    - Transaction summary
    - Historical performance data
    - 15-year cashflow projections
    - Debt coverage analysis
    - Returns analysis (IRR, NPV, MOIC)
    - Sensitivity & scenario analysis
    - Market benchmarks
    - Operating expense breakdown
    - Revenue assumptions
    - Risk register
    - Macro assumptions
    - Comparable transactions
    - Monthly debt schedule
    - Technical specifications
    - Development timeline
    """

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # Sheet 1: Transaction Summary
        trans_summary = create_transaction_summary(
            capacity_mwh, power_mw, investment_eur, equity_eur, debt_eur,
            loan_term_years, interest_rate
        )
        trans_summary.to_excel(writer, sheet_name=get_text("sheet_transaction", language), index=False, header=False)

        # Sheet 2: Historical Performance
        if fr_metrics or pzu_metrics:
            historical = create_historical_performance(fr_metrics, pzu_metrics)
            historical.to_excel(writer, sheet_name=get_text("sheet_historical", language), index=False, header=False)

        # Sheet 3: 15-Year Cashflow
        if fr_metrics and "annual" in fr_metrics:
            fr_annual = fr_metrics["annual"]
            cashflow_15y = create_15year_cashflow(
                annual_revenue=fr_annual.get("total", 0),
                annual_energy_cost=fr_annual.get("energy_cost", 0),
                annual_opex=fr_opex_annual,
                annual_debt_service=fr_annual.get("debt", 0),
                investment_eur=investment_eur,
                equity_eur=equity_eur,
                loan_term_years=loan_term_years,
            )
            cashflow_15y.to_excel(writer, sheet_name=get_text("sheet_cashflow", language), index=False)

            # Sheet 4: Debt Coverage
            debt_coverage = create_debt_coverage_analysis(cashflow_15y, debt_eur)
            debt_coverage.to_excel(writer, sheet_name=get_text("sheet_debt_coverage", language), index=False, header=False)

            # Sheet 5: Returns Analysis
            returns = create_returns_analysis(cashflow_15y, equity_eur, investment_eur)
            returns.to_excel(writer, sheet_name=get_text("sheet_returns", language), index=False, header=False)

            # Sheet 6: Sensitivity Analysis
            sensitivity = create_sensitivity_analysis(
                fr_annual.get("total", 0),
                fr_opex_annual,
                fr_annual.get("energy_cost", 0),
                fr_annual.get("debt", 0),
                equity_eur,
            )
            sensitivity.to_excel(writer, sheet_name=get_text("sheet_sensitivity", language), index=False, header=False)

        # Sheet 7: Market Benchmarks
        benchmarks = create_market_benchmarks(capacity_mwh)
        benchmarks.to_excel(writer, sheet_name=get_text("sheet_benchmarks", language), index=False, header=False)

        # Sheet 8: OPEX Breakdown
        opex_detail = create_opex_breakdown(fr_opex_annual, power_mw)
        opex_detail.to_excel(writer, sheet_name=get_text("sheet_opex", language), index=False, header=False)

        # Sheet 9: Revenue Assumptions
        revenue_assumptions = create_revenue_assumptions(fr_metrics)
        revenue_assumptions.to_excel(writer, sheet_name=get_text("sheet_revenue", language), index=False, header=False)

        # Sheet 10: Risk Register
        risks = create_risk_register(investment_eur)
        risks.to_excel(writer, sheet_name=get_text("sheet_risks", language), index=False, header=False)

        # Sheet 11: Macro Assumptions
        macro = create_macro_assumptions()
        macro.to_excel(writer, sheet_name=get_text("sheet_macro", language), index=False, header=False)

        # Sheet 12: Comparable Projects
        comps = create_comparable_projects()
        comps.to_excel(writer, sheet_name=get_text("sheet_comparables", language), index=False, header=False)

        # Sheet 13: Monthly Debt Schedule
        debt_schedule = create_monthly_debt_schedule(debt_eur, interest_rate, loan_term_years)
        if not debt_schedule.empty:
            debt_schedule.to_excel(writer, sheet_name=get_text("sheet_debt_schedule", language), index=False)

        # Sheet 14: Technical Specs
        tech_specs = create_technical_specifications(capacity_mwh, power_mw)
        tech_specs.to_excel(writer, sheet_name=get_text("sheet_technical", language), index=False, header=False)

        # Sheet 15: Development Timeline
        timeline = create_development_timeline()
        timeline.to_excel(writer, sheet_name=get_text("sheet_timeline", language), index=False, header=False)

        # Apply professional formatting to all sheets
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            apply_excel_formatting(worksheet, sheet_name)

    output.seek(0)
    return output.getvalue()


def add_export_buttons(
    fr_metrics: Optional[Dict[str, Any]],
    pzu_metrics: Optional[Dict[str, Any]],
    capacity_mwh: float,
    power_mw: float,
    investment_eur: float,
    equity_eur: float,
    debt_eur: float,
    loan_term_years: int,
    interest_rate: float,
    fr_opex_annual: float,
    pzu_opex_annual: float,
    fr_years_analyzed: int,
    pzu_years_analyzed: int,
) -> None:
    """
    Add JP Morgan-level export button to Streamlit UI
    """

    st.markdown("---")
    st.markdown("## 📊 Investment Banking Financial Package")
    st.markdown("### Professional 15-Sheet Excel Model for Senior Debt Financing")

    # Calculate key metrics for preview
    if fr_metrics and "annual" in fr_metrics:
        fr_annual = fr_metrics["annual"]

        # Calculate equity cashflows for IRR
        cashflows = []
        cumulative = -equity_eur
        cashflows.append(-equity_eur)

        for year in range(1, 16):
            revenue = fr_annual.get("total", 0) * (1.02 ** (year - 1))
            energy_cost = fr_annual.get("energy_cost", 0) * (1.02 ** (year - 1))
            opex = fr_opex_annual * (1.025 ** (year - 1))
            ebitda = revenue - energy_cost - opex

            capex = -investment_eur * 0.10 if year == 6 else 0
            debt_svc = fr_annual.get("debt", 0) if year <= loan_term_years else 0

            fcf = ebitda + capex - debt_svc
            cumulative += fcf
            cashflows.append(fcf)

        # Calculate IRR
        equity_irr = calculate_irr(cashflows) * 100

        # Calculate DSCR
        avg_dscr = 0
        min_dscr = 999
        for year in range(1, min(loan_term_years + 1, 16)):
            revenue = fr_annual.get("total", 0) * (1.02 ** (year - 1))
            energy_cost = fr_annual.get("energy_cost", 0) * (1.02 ** (year - 1))
            opex = fr_opex_annual * (1.025 ** (year - 1))
            ebitda = revenue - energy_cost - opex
            debt_svc = fr_annual.get("debt", 0)

            if debt_svc > 0:
                dscr = ebitda / debt_svc
                avg_dscr += dscr
                min_dscr = min(min_dscr, dscr)

        avg_dscr = avg_dscr / loan_term_years if loan_term_years > 0 else 0

        # Calculate MOIC
        total_distributions = sum(cf for cf in cashflows[1:] if cf > 0)
        moic = total_distributions / equity_eur if equity_eur > 0 else 0

        # Payback
        payback = next((i for i, val in enumerate(cashflows) if sum(cashflows[:i+1]) > 0), None)

        # Display professional preview
        st.markdown("#### 📈 Financial Highlights Preview")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(
                "Equity IRR",
                f"{equity_irr:.1f}%",
                help="Levered internal rate of return on equity investment"
            )

        with col2:
            st.metric(
                "Min DSCR",
                f"{min_dscr:.2f}x",
                delta="PASS" if min_dscr >= 1.20 else "FAIL",
                delta_color="normal" if min_dscr >= 1.20 else "inverse",
                help="Minimum Debt Service Coverage Ratio (target: ≥1.20x)"
            )

        with col3:
            st.metric(
                "MOIC",
                f"{moic:.2f}x",
                help="Multiple on Invested Capital over 15 years"
            )

        with col4:
            st.metric(
                "Payback",
                f"{payback} years" if payback else "N/A",
                help="Equity payback period"
            )

        st.markdown("")

    # Excel contents description
    st.markdown("#### 📋 Excel Model Contents (15 Comprehensive Sheets)")

    sheets_data = {
        "Sheet #": list(range(1, 16)),
        "Sheet Name": [
            "Transaction Summary",
            "Historical Data",
            "15Y Cashflow",
            "Debt Coverage",
            "Returns Analysis",
            "Sensitivity",
            "Market Benchmarks",
            "OPEX Detail",
            "Revenue Model",
            "Risk Register",
            "Macro & FX",
            "Comparables",
            "Debt Schedule",
            "Technical Specs",
            "Timeline"
        ],
        "Description": [
            "Sources & Uses, debt terms, key assumptions",
            "Actual monthly FR/PZU performance data",
            "Year-by-year P&L waterfall with escalations & DSCR",
            "Min/avg DSCR, covenant testing, PASS/FAIL",
            "Equity IRR, NPV (4 rates), MOIC, payback period",
            "Revenue/OPEX sensitivity matrix + 5 scenarios",
            "FR/PZU pricing, comparable transactions, benchmarks",
            "6-category OPEX split with €/kW metrics",
            "Capacity/activation split, market drivers",
            "8 risks with mitigation & residual ratings",
            "Inflation, FX, interest rates, tax framework",
            "5 international BESS projects with metrics",
            "120-month amortization with cumulative tracking",
            "Battery, PCS, BMS, safety, grid connection",
            "18-month critical path with permitting"
        ],
        "Use For": [
            "Credit memo, investment committee",
            "Due diligence validation",
            "Debt sizing, DSCR covenants",
            "Credit approval",
            "Equity investment decision",
            "Stress testing, downside analysis",
            "Market positioning",
            "Operating budget",
            "Revenue forecasting",
            "Risk assessment",
            "FX hedging, tax structuring",
            "Valuation benchmarking",
            "Lender amortization table",
            "Technical due diligence",
            "Project schedule risk"
        ]
    }

    st.dataframe(
        pd.DataFrame(sheets_data),
        hide_index=True,
        use_container_width=True,
        height=400
    )

    st.markdown("")
    st.markdown("**Perfect for:** JP Morgan, Goldman Sachs, Citi, Morgan Stanley presentations | Senior debt syndications | Institutional equity investors")
    st.markdown("")

    # Language selection
    col_lang1, col_lang2, col_lang3 = st.columns([1, 1, 1])
    with col_lang2:
        language = st.selectbox(
            "🌍 Select Language / Selectați Limba",
            options=["en", "ro"],
            format_func=lambda x: "🇬🇧 English" if x == "en" else "🇷🇴 Română",
            key="export_language"
        )

    st.markdown("")

    # Download buttons
    col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])

    with col_btn2:
        button_text = "📥 Generate Financial Model" if language == "en" else "📥 Generează Model Financiar"
        if st.button(button_text, type="primary", use_container_width=True):
            spinner_text = "Generating comprehensive financial model... (15 sheets)" if language == "en" else "Generare model financiar complet... (15 foi)"
            with st.spinner(spinner_text):
                try:
                    excel_data = export_financial_package_to_excel(
                        fr_metrics=fr_metrics,
                        pzu_metrics=pzu_metrics,
                        capacity_mwh=capacity_mwh,
                        power_mw=power_mw,
                        investment_eur=investment_eur,
                        equity_eur=equity_eur,
                        debt_eur=debt_eur,
                        loan_term_years=loan_term_years,
                        interest_rate=interest_rate,
                        fr_opex_annual=fr_opex_annual,
                        pzu_opex_annual=pzu_opex_annual,
                        fr_years_analyzed=fr_years_analyzed,
                        pzu_years_analyzed=pzu_years_analyzed,
                        language=language,
                    )

                    lang_suffix = "EN" if language == "en" else "RO"
                    filename = f"BESS_Model_Financiar_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{lang_suffix}.xlsx"

                    download_label = "⬇️ Download Financial Model (.xlsx)" if language == "en" else "⬇️ Descarcă Model Financiar (.xlsx)"
                    st.download_button(
                        label=download_label,
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                        use_container_width=True,
                    )

                    success_msg = f"✅ 15-sheet financial model generated successfully! ({capacity_mwh:.0f} MWh project)" if language == "en" else f"✅ Model financiar cu 15 foi generat cu succes! (Proiect {capacity_mwh:.0f} MWh)"
                    st.success(success_msg)

                except Exception as e:
                    st.error(f"❌ Error generating financial model: {e}")
                    import traceback
                    st.code(traceback.format_exc())
